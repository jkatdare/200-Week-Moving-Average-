import io
import pandas as pd
import yfinance as yf
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import requests
from bs4 import BeautifulSoup

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="200 Week Moving Average Scanner",
    page_icon="📈",
    layout="wide",
)

# ── Styling ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    [data-testid="stAppViewContainer"] { background: #0e1117; }
    [data-testid="stSidebar"] { background: #161b22; border-right: 1px solid #30363d; }
    h1 { color: #58a6ff !important; letter-spacing: -1px; }
    .metric-card {
        background: #161b22;
        border: 1px solid #30363d;
        border-radius: 8px;
        padding: 16px 20px;
        text-align: center;
    }
    .above { color: #3fb950; font-weight: 600; }
    .below { color: #f85149; font-weight: 600; }
    div[data-testid="stDataFrame"] { border: 1px solid #30363d; border-radius: 8px; }
</style>
""", unsafe_allow_html=True)

WEEKS = 200
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
SOURCES = {
    "SP500":     "https://slickcharts.com/sp500",
    "NASDAQ100": "https://slickcharts.com/nasdaq100",
    "DOW30":     "https://slickcharts.com/dowjones",
}

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Configuration")
    st.markdown("---")

    selected_indices = st.multiselect(
        "Indices to scan",
        options=list(SOURCES.keys()),
        default=list(SOURCES.keys()),
    )

    max_above_pct = st.slider(
        "Max % above 200 WMA",
        min_value=-50,
        max_value=50,
        value=10,
        step=1,
        help="Only show stocks within this % above (or any % below) the 200 WMA",
    )

    st.markdown("---")
    run_btn = st.button("🚀 Run Analysis", use_container_width=True, type="primary")
    st.markdown("---")
    st.markdown("**About**")
    st.caption(
        "Scans S&P 500, NASDAQ 100, and DJIA 30 for stocks "
        "trading near or below their 200-week moving average."
    )

# ── Header ────────────────────────────────────────────────────────────────────
st.title("📈 200 Week Moving Average Scanner")
st.caption(f"Identifies stocks trading at or below {max_above_pct}% above their 200-week MA")

# ── Helper functions ──────────────────────────────────────────────────────────
@st.cache_data(ttl=3600, show_spinner=False)
def scrape_tickers(url):
    html = requests.get(url, headers=HEADERS, timeout=15).text
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    tickers = []
    for row in table.find_all("tr")[1:]:
        cols = row.find_all("td")
        if len(cols) >= 3:
            ticker = cols[2].text.strip().replace(".", "-")
            if ticker:
                tickers.append(ticker)
    return tickers


def get_tickers(indices):
    all_tickers = []
    for idx in indices:
        tickers = scrape_tickers(SOURCES[idx])
        all_tickers.extend(tickers)
    return list(dict.fromkeys(all_tickers))


@st.cache_data(ttl=3600, show_spinner=False)
def get_price_data(tickers_tuple):
    tickers = list(tickers_tuple)
    start = (datetime.today() - timedelta(weeks=WEEKS + 60)).strftime("%Y-%m-%d")
    end = datetime.today().strftime("%Y-%m-%d")
    raw = yf.download(tickers, start=start, end=end,
                      interval="1wk", auto_adjust=True, progress=False)
    close = raw["Close"].dropna(axis=1, thresh=WEEKS)
    return close


def build_results(close, max_above):
    current_price = close.iloc[-1]
    wma_200 = close.rolling(window=WEEKS).mean().iloc[-1]
    results = pd.DataFrame({
        "Current Price": current_price.round(2),
        "200 WMA": wma_200.round(2),
    })
    results["% vs 200 WMA"] = (
        (results["Current Price"] - results["200 WMA"]) / results["200 WMA"] * 100
    ).round(2)
    filtered = results[results["% vs 200 WMA"] <= max_above].copy()
    filtered["Signal"] = filtered["% vs 200 WMA"].apply(lambda x: "ABOVE" if x >= 0 else "BELOW")
    filtered.sort_values("% vs 200 WMA", ascending=False, inplace=True)
    filtered.index.name = "Ticker"
    return filtered


@st.cache_data(ttl=3600, show_spinner=False)
def get_company_names(tickers_tuple):
    names = {}
    for ticker in tickers_tuple:
        try:
            names[ticker] = yf.Ticker(ticker).info.get("longName", ticker)
        except Exception:
            names[ticker] = ticker
    return names


def make_chart_image(ticker, close):
    series = close[ticker].dropna().iloc[-WEEKS:]
    wma = close[ticker].rolling(window=WEEKS).mean().dropna().iloc[-WEEKS:]
    fig, ax = plt.subplots(figsize=(6, 2.5), facecolor="#0e1117")
    ax.set_facecolor("#0e1117")
    ax.plot(series.index, series.values, label="Price", color="#58a6ff", linewidth=1.2)
    ax.plot(wma.index, wma.values, label="200 WMA", color="#f0883e", linewidth=1.2, linestyle="--")
    ax.set_title(ticker, fontsize=9, fontweight="bold", color="white")
    ax.legend(fontsize=7, facecolor="#161b22", labelcolor="white")
    ax.tick_params(labelsize=7, colors="gray")
    for spine in ax.spines.values():
        spine.set_edgecolor("#30363d")
    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=100, facecolor="#0e1117")
    plt.close(fig)
    buf.seek(0)
    return buf


def build_excel(results, close, names):
    wb = Workbook()
    ws = wb.active
    ws.title = "200 WMA Results"
    headers = ["Ticker", "Company", "Current Price", "200 WMA", "% vs 200 WMA", "Signal", "Chart"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 65
    ROW_HEIGHT = 95
    for i, (ticker, row) in enumerate(results.iterrows(), start=2):
        ws.row_dimensions[i].height = ROW_HEIGHT
        ws.cell(row=i, column=1, value=ticker).font = Font(bold=True, name="Arial")
        ws.cell(row=i, column=2, value=names.get(ticker, ticker))
        ws.cell(row=i, column=3, value=row["Current Price"])
        ws.cell(row=i, column=4, value=row["200 WMA"])
        pct_cell = ws.cell(row=i, column=5, value=row["% vs 200 WMA"] / 100)
        pct_cell.number_format = "0.00%"
        pct_cell.font = Font(
            color="375623" if row["% vs 200 WMA"] >= 0 else "9C0006",
            name="Arial"
        )
        ws.cell(row=i, column=6, value=row["Signal"])
        img_buf = make_chart_image(ticker, close)
        xl_img = XLImage(img_buf)
        xl_img.width = 450
        xl_img.height = 120
        ws.add_image(xl_img, f"G{i}")
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ── Main analysis ─────────────────────────────────────────────────────────────
if run_btn:
    if not selected_indices:
        st.warning("Please select at least one index.")
        st.stop()

    with st.status("Running analysis...", expanded=True) as status:
        st.write(f"📋 Scraping tickers for: {', '.join(selected_indices)}")
        tickers = get_tickers(selected_indices)
        st.write(f"✅ Found **{len(tickers)}** unique tickers")

        st.write("📥 Downloading weekly price data (this may take a minute)...")
        close = get_price_data(tuple(tickers))
        st.write(f"✅ Price data loaded for **{len(close.columns)}** tickers")

        st.write("🔢 Calculating 200 WMA and filtering results...")
        results = build_results(close, max_above_pct)
        st.write(f"✅ **{len(results)}** stocks match your criteria")

        status.update(label="Analysis complete!", state="complete")

    # ── Metrics ───────────────────────────────────────────────────────────────
    col1, col2, col3, col4 = st.columns(4)
    above = (results["Signal"] == "ABOVE").sum()
    below = (results["Signal"] == "BELOW").sum()
    with col1:
        st.metric("Stocks Found", len(results))
    with col2:
        st.metric("Above 200 WMA", above)
    with col3:
        st.metric("Below 200 WMA", below)
    with col4:
        avg_pct = results["% vs 200 WMA"].mean()
        st.metric("Avg % vs 200 WMA", f"{avg_pct:.1f}%")

    st.markdown("---")

    # ── Table ─────────────────────────────────────────────────────────────────
    st.subheader("📊 Results")

    def color_signal(val):
        return "color: #3fb950; font-weight:600" if val == "ABOVE" else "color: #f85149; font-weight:600"

    def color_pct(val):
        return "color: #3fb950" if val >= 0 else "color: #f85149"

    styled = results.style \
        .map(color_signal, subset=["Signal"]) \
        .map(color_pct, subset=["% vs 200 WMA"]) \
        .format({"Current Price": "${:.2f}", "200 WMA": "${:.2f}", "% vs 200 WMA": "{:.2f}%"})

    st.dataframe(styled, use_container_width=True, height=500)

    st.markdown("---")

    # ── Charts ────────────────────────────────────────────────────────────────
    st.subheader("📉 Price Charts")
    top_n = min(20, len(results))
    st.caption(f"Showing top {top_n} results")

    cols_per_row = 3
    tickers_to_show = results.index[:top_n].tolist()
    for i in range(0, len(tickers_to_show), cols_per_row):
        cols = st.columns(cols_per_row)
        for j, ticker in enumerate(tickers_to_show[i:i + cols_per_row]):
            with cols[j]:
                series = close[ticker].dropna().iloc[-WEEKS:]
                wma = close[ticker].rolling(window=WEEKS).mean().dropna().iloc[-WEEKS:]
                fig, ax = plt.subplots(figsize=(4, 2), facecolor="#161b22")
                ax.set_facecolor("#161b22")
                ax.plot(series.index, series.values, color="#58a6ff", linewidth=1.1)
                ax.plot(wma.index, wma.values, color="#f0883e", linewidth=1.1, linestyle="--")
                pct = results.loc[ticker, "% vs 200 WMA"]
                color = "#3fb950" if pct >= 0 else "#f85149"
                ax.set_title(f"{ticker}  ({pct:+.1f}%)", fontsize=8, color=color, fontweight="bold")
                ax.tick_params(labelsize=6, colors="#8b949e")
                for spine in ax.spines.values():
                    spine.set_edgecolor("#30363d")
                plt.tight_layout()
                st.pyplot(fig, use_container_width=True)
                plt.close(fig)

    st.markdown("---")

    # ── Excel download ────────────────────────────────────────────────────────
    st.subheader("⬇️ Download Results")
    with st.spinner("Building Excel file with charts..."):
        names = get_company_names(tuple(results.index.tolist()))
        excel_bytes = build_excel(results, close, names)

    fname = f"200wma_results_{datetime.today().strftime('%Y%m%d')}.xlsx"
    st.download_button(
        label="📥 Download Excel Report",
        data=excel_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

else:
    st.info("👈 Configure your settings in the sidebar and click **Run Analysis** to get started.")
    st.markdown("""
    ### How it works
    1. **Select indices** — choose from S&P 500, NASDAQ 100, and/or DJIA 30
    2. **Set threshold** — the max % above the 200-week MA to include
    3. **Run** — live data is fetched from Yahoo Finance
    4. **Download** — get a formatted Excel report with mini-charts

    > The 200-week moving average is a long-term trend indicator. Stocks trading at or below it 
    > are often considered historically cheap on a long-term basis.
    """)
