import io
import os
import pandas as pd
import yfinance as yf
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
from tickers import get_tickers
from config import MAX_ABOVE_PCT

WEEKS       = 200
START_DATE  = (datetime.today() - timedelta(weeks=WEEKS + 60)).strftime("%Y-%m-%d")
END_DATE    = datetime.today().strftime("%Y-%m-%d")
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "200wma_results.xlsx")


def get_data(tickers):
    print(f"Downloading price data for {len(tickers)} tickers...")
    raw = yf.download(tickers, start=START_DATE, end=END_DATE,
                      interval="1wk", auto_adjust=True, progress=False)
    close = raw["Close"].dropna(axis=1, thresh=WEEKS)
    return close


def build_results(close):
    current_price = close.iloc[-1]
    wma_200       = close.rolling(window=WEEKS).mean().iloc[-1]

    results = pd.DataFrame({
        "Current Price": current_price.round(2),
        "200 WMA":       wma_200.round(2),
    })
    results["% vs 200 WMA"] = (
        (results["Current Price"] - results["200 WMA"]) / results["200 WMA"] * 100
    ).round(2)

    filtered = results[results["% vs 200 WMA"] <= MAX_ABOVE_PCT].copy()
    filtered["Signal"] = filtered["% vs 200 WMA"].apply(lambda x: "ABOVE" if x >= 0 else "BELOW")
    filtered.sort_values("% vs 200 WMA", ascending=False, inplace=True)
    filtered.index.name = "Ticker"
    return filtered


def get_company_names(tickers):
    print(f"Fetching company names for {len(tickers)} stocks...")
    names = {}
    for ticker in tickers:
        try:
            names[ticker] = yf.Ticker(ticker).info.get("longName", ticker)
        except:
            names[ticker] = ticker
    return names


def make_chart_image(ticker, close):
    series = close[ticker].dropna().iloc[-WEEKS:]
    wma    = close[ticker].rolling(window=WEEKS).mean().dropna().iloc[-WEEKS:]

    fig, ax = plt.subplots(figsize=(6, 2.5))
    ax.plot(series.index, series.values, label="Price", color="#1f77b4", linewidth=1.2)
    ax.plot(wma.index, wma.values, label="200 WMA", color="#ff7f0e", linewidth=1.2, linestyle="--")
    ax.set_title(ticker, fontsize=9, fontweight="bold")
    ax.legend(fontsize=7)
    ax.tick_params(labelsize=7)
    plt.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=100)
    plt.close(fig)
    buf.seek(0)
    return buf


def save_to_excel(results, close):
    names = get_company_names(results.index.tolist())

    wb = Workbook()
    ws = wb.active
    ws.title = "200 WMA Results"

    headers = ["Ticker", "Company", "Current Price", "200 WMA", "% vs 200 WMA", "Signal", "Chart"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill      = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 65

    ROW_HEIGHT = 95

    for i, (ticker, row) in enumerate(results.iterrows(), start=2):
        ws.row_dimensions[i].height = ROW_HEIGHT

        ws.cell(row=i, column=1, value=ticker).font = Font(bold=True, name="Arial")
        ws.cell(row=i, column=2, value=names.get(ticker, ticker))
        ws.cell(row=i, column=3, value=row["Current Price"])
        ws.cell(row=i, column=4, value=row["200 WMA"])

        pct_cell = ws.cell(row=i, column=5, value=row["% vs 200 WMA"] / 100)
        pct_cell.number_format = "0.00%"
        pct_cell.font = Font(
            color="375623" if row["% vs 200 WMA"] >= 0 else "9C0006",
            name="Arial"
        )

        ws.cell(row=i, column=6, value=row["Signal"])

        img_buf = make_chart_image(ticker, close)
        xl_img  = XLImage(img_buf)
        xl_img.width  = 450
        xl_img.height = 120
        ws.add_image(xl_img, f"G{i}")

    wb.save(OUTPUT_FILE)
    print(f"\nDone! File saved to:\n{OUTPUT_FILE}")


def main():
    tickers = get_tickers()
    close   = get_data(tickers)
    results = build_results(close)
    save_to_excel(results, close)


if __name__ == "__main__":
    main()
